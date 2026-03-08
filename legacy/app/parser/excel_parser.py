"""
Format-agnostic Excel parser.

Strategy
--------
1. Auto-detect the sheet structure (header row, month columns, label column,
   account-code column) using heuristics — no hard-coded row/column numbers.
2. Walk every row and classify it as: header | detail line | subtotal/total.
3. Fuzzy-match known financial labels to build the `key_figures` dict.
4. Return a `FinancialStatement` regardless of how the source sheet is laid out.
"""

import re
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from app.models.statement import (
    FinancialStatement,
    LineItem,
    SheetStructure,
)

# ── Pattern constants ──────────────────────────────────────────────────────────

MONTH_ABBREVS = ["jan", "feb", "mar", "apr", "may", "jun",
                 "jul", "aug", "sep", "oct", "nov", "dec"]
MONTH_FULL    = ["january", "february", "march", "april", "may", "june",
                 "july", "august", "september", "october", "november", "december"]

MONTH_RE     = re.compile(
    r"\b(" + "|".join(MONTH_ABBREVS + MONTH_FULL) + r")\b",
    re.IGNORECASE,
)
QUARTER_RE   = re.compile(r"\bQ[1-4]\b", re.IGNORECASE)
YM_RE        = re.compile(r"\b\d{1,2}/\d{4}\b")          # 01/2025
YEAR_RE      = re.compile(r"\b(20\d{2})\b")
ACCT_CODE_RE = re.compile(r"^\d{3,5}-\d{3,5}$")           # 5120-0000 style
SUBTOTAL_RE  = re.compile(
    r"^\s*(total|net|cash flow|gross income|grand total)", re.IGNORECASE
)

# Fuzzy keyword map: semantic_name -> list of substrings to search for (any match wins)
KEY_FIGURE_PATTERNS: Dict[str, List[str]] = {
    "gross_potential_rent":      ["total gross potential rent", "gross potential rent"],
    "vacancy_loss":              ["loss due to vacancies", "vacancy loss"],
    "concession_loss":           ["loss due to concessions", "concession loss"],
    "office_model_rent_free":    ["total office model & rent free", "office model & rent free", "office model"],
    "bad_debt":                  ["net bad debt expense (recovery)", "net bad debt expense", "net bad debt"],
    "net_rental_revenue":        ["net rental revenue"],
    "other_tenant_charges":      ["total other tenant charges"],
    "total_revenue":             ["total revenue"],
    "controllable_expenses":     ["total controllable expenses"],
    "non_controllable_expenses": ["total non-controllable expenses"],
    "total_operating_expenses":  ["total operating expenses"],
    "noi":                       ["net operating income (loss)", "net operating income"],
    "total_payroll":             ["total payroll and benefits"],
    "management_fees":           ["total management fees"],
    "utilities":                 ["total utilities expense", "total utilities"],
    "real_estate_taxes":         ["total real estate taxes"],
    "insurance":                 ["total property and liability insurance"],
    "financial_expense":         ["total financial expense"],
    "replacement_expense":       ["total replacement expense"],
    "total_non_operating":       ["total non-operating expense (revenue)", "total non-operating expense"],
    "net_income":                ["net income (loss)", "net income"],
    "cash_flow":                 ["cash flow"],
}


# ── Public API ─────────────────────────────────────────────────────────────────

def parse_excel(path: str, sheet_name: Optional[str] = None) -> FinancialStatement:
    """
    Parse an Excel file and return a FinancialStatement.

    Parameters
    ----------
    path        : path to the .xlsx file
    sheet_name  : sheet to parse; if None the first/best sheet is auto-detected
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _pick_sheet(wb, sheet_name)
    raw = _read_raw(ws)
    structure = _detect_structure(raw)
    rows = _parse_rows(raw, structure)
    key_figures = _extract_key_figures(rows)
    meta = _extract_metadata(raw, structure)

    return FinancialStatement(
        property_name=meta.get("property_name", "Unknown Property"),
        period=meta.get("period", "Unknown Period"),
        book_type=meta.get("book_type", ""),
        months=[m for _, m in structure.month_cols],
        all_rows=rows,
        key_figures=key_figures,
        structure=structure,
        raw_data=raw,
    )


# ── Internal helpers ───────────────────────────────────────────────────────────

def _pick_sheet(wb: openpyxl.Workbook, name: Optional[str]):
    if name and name in wb.sheetnames:
        return wb[name]
    # Prefer sheets whose name contains financial keywords
    for sn in wb.sheetnames:
        low = sn.lower()
        if any(k in low for k in ("report", "p&l", "income", "statement", "pl")):
            return wb[sn]
    return wb[wb.sheetnames[0]]


def _read_raw(ws) -> List[List[Any]]:
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows()
    ]


def _detect_structure(raw: List[List[Any]]) -> SheetStructure:
    """
    Heuristically determine:
    - Which row contains the column headers (months)
    - Which columns hold the month data, totals, account codes, and labels
    """
    header_row, month_cols = _find_header_row(raw)
    if header_row is None:
        raise ValueError(
            "Could not detect a header row with month names. "
            "Please ensure the spreadsheet has monthly column headers."
        )

    total_col = _find_total_col(raw[header_row])
    account_col, label_col = _find_label_cols(raw, header_row + 1)
    data_start_row = header_row + 1

    return SheetStructure(
        header_row=header_row,
        month_cols=month_cols,
        total_col=total_col,
        account_col=account_col,
        label_col=label_col,
        data_start_row=data_start_row,
    )


def _find_header_row(raw: List[List[Any]]) -> Tuple[Optional[int], List[Tuple[int, str]]]:
    """Scan the first 30 rows for a row that contains ≥3 month-like values."""
    for row_idx, row in enumerate(raw[:30]):
        hits: List[Tuple[int, str]] = []
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if MONTH_RE.search(cell_str) or QUARTER_RE.search(cell_str) or YM_RE.search(cell_str):
                hits.append((col_idx, cell_str))
        if len(hits) >= 3:
            return row_idx, hits
    return None, []


def _find_total_col(header_row: List[Any]) -> Optional[int]:
    for col_idx, cell in enumerate(header_row):
        if cell and "total" in str(cell).lower():
            return col_idx
    return None


def _find_label_cols(
    raw: List[List[Any]], start_row: int
) -> Tuple[Optional[int], int]:
    """
    Scan data rows for account-code patterns to locate the account and label columns.
    Falls back to the widest-text column if no account codes exist.
    """
    account_col: Optional[int] = None
    label_col: int = 0

    for row in raw[start_row: start_row + 40]:
        for col_idx, cell in enumerate(row):
            if cell and ACCT_CODE_RE.match(str(cell).strip()):
                account_col = col_idx
                # Label is typically the next column
                label_col = col_idx + 1
                return account_col, label_col

    # No account codes — find the column with the most non-empty text cells
    col_text_count: Dict[int, int] = {}
    for row in raw[start_row: start_row + 40]:
        for col_idx, cell in enumerate(row):
            if cell and isinstance(cell, str) and len(cell.strip()) > 3:
                col_text_count[col_idx] = col_text_count.get(col_idx, 0) + 1
    if col_text_count:
        label_col = max(col_text_count, key=col_text_count.get)

    return None, label_col


def _parse_rows(raw: List[List[Any]], s: SheetStructure) -> List[LineItem]:
    """Convert raw cell data into a list of LineItem objects."""
    month_indices = [ci for ci, _ in s.month_cols]
    month_names   = [mn for _, mn in s.month_cols]
    rows: List[LineItem] = []

    for row_idx, row in enumerate(raw[s.data_start_row:], start=s.data_start_row):
        # Pad short rows
        row = list(row) + [None] * (max(month_indices + [s.label_col + 1]) - len(row) + 1)

        label_cell = row[s.label_col] if s.label_col < len(row) else None
        if label_cell is None or str(label_cell).strip() == "":
            continue

        label = str(label_cell).strip()
        account_code = None
        if s.account_col is not None and s.account_col < len(row):
            ac = row[s.account_col]
            if ac and ACCT_CODE_RE.match(str(ac).strip()):
                account_code = str(ac).strip()

        monthly_values: Dict[str, Optional[float]] = {}
        for col_idx, month_name in zip(month_indices, month_names):
            monthly_values[month_name] = _to_float(row[col_idx] if col_idx < len(row) else None)

        annual_total = None
        if s.total_col is not None and s.total_col < len(row):
            annual_total = _to_float(row[s.total_col])

        indent = _measure_indent(label_cell)
        is_subtotal = bool(SUBTOTAL_RE.match(label))
        is_header = (account_code is None) and (not is_subtotal) and (not any(monthly_values.values()))

        rows.append(LineItem(
            label=label,
            monthly_values=monthly_values,
            annual_total=annual_total,
            row_number=row_idx + 1,        # 1-based for display
            account_code=account_code,
            is_subtotal=is_subtotal,
            is_header=is_header,
            indent_level=indent,
        ))

    return rows


def _extract_key_figures(rows: List[LineItem]) -> Dict[str, LineItem]:
    """
    Match each semantic key figure to the best matching LineItem.

    Scoring priority (highest wins):
      1. Exact label match with the pattern          (+100)
      2. Pattern covers >80% of the label length     (+50)
      3. Row has actual annual or monthly values      (+20)
      4. Row is a subtotal                            (+15)
      5. Pattern is any substring of label            (+10)
      Penalty: header row with no values              (-30)
    """
    key_figures: Dict[str, LineItem] = {}

    for semantic_name, patterns in KEY_FIGURE_PATTERNS.items():
        best: Optional[LineItem] = None
        best_score: int = -1

        for item in rows:
            label_lower = item.label.lower().strip()
            for pattern in patterns:
                if pattern not in label_lower:
                    continue
                score = 10  # base: substring matched

                # Prefer near-exact matches
                if label_lower == pattern:
                    score += 100
                elif len(label_lower) > 0 and len(pattern) / len(label_lower) > 0.8:
                    score += 50

                # Prefer rows with real values
                if item.annual_total is not None or item.has_any_value():
                    score += 20

                # Prefer subtotals
                if item.is_subtotal:
                    score += 15

                # Penalise headers with no data
                if item.is_header:
                    score -= 30

                if score > best_score:
                    best_score = score
                    best = item
                break  # only score once per row (first matching pattern wins)

        if best is not None:
            key_figures[semantic_name] = best

    return key_figures


def _extract_metadata(raw: List[List[Any]], s: SheetStructure) -> Dict[str, str]:
    """Pull property name, period, and book type from the metadata rows above the header."""
    meta: Dict[str, str] = {}
    for row in raw[: s.header_row]:
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            # Skip spreadsheet system artifacts like "Tree =", "Page =", "View ="
            if re.match(r'^\w[\w\s]{0,20}=\s*$', text):
                continue
            # First substantial non-empty text = property name
            if "property_name" not in meta and len(text) > 3:
                # Strip trailing system codes like ".secesml" (dot + lowercase, no spaces)
                cleaned = re.sub(r"\.[a-z]{2,20}$", "", text).strip()
                meta["property_name"] = cleaned if len(cleaned) > 3 else text
            # Period detection
            if "period" not in meta and ("period" in text.lower() or YEAR_RE.search(text)):
                meta["period"] = text
            # Book type
            if "book_type" not in meta and "book" in text.lower():
                meta["book_type"] = text
    return meta


def enrich_key_figures(
    stmt,
    label_map: Dict[str, str],
) -> None:
    """
    Fill gaps in stmt.key_figures using a mapping returned by LabelMapperAgent.

    label_map: {semantic_name: row_label_as_identified_by_llm}

    Mutates stmt.key_figures in-place.  Concepts already present are not
    overwritten — heuristic matches take precedence.
    """
    for concept, target_label in label_map.items():
        if concept in stmt.key_figures:
            continue    # heuristic already found it
        target_lower = target_label.lower().strip()
        best = None
        best_score = -1
        for item in stmt.all_rows:
            label_lower = item.label.lower().strip()
            if label_lower != target_lower and target_lower not in label_lower:
                continue
            score = 10
            if label_lower == target_lower:
                score += 100
            elif len(label_lower) > 0 and len(target_lower) / len(label_lower) > 0.8:
                score += 50
            if item.has_any_value() or item.annual_total is not None:
                score += 20
            if item.is_subtotal:
                score += 15
            if item.is_header:
                score -= 30
            if score > best_score:
                best_score = score
                best = item
        if best is not None:
            stmt.key_figures[concept] = best


def _to_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("$", "").replace("(", "-").replace(")", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _measure_indent(raw_label: Any) -> int:
    """Count leading spaces to infer hierarchy depth."""
    if raw_label is None:
        return 0
    s = str(raw_label)
    leading = len(s) - len(s.lstrip(" "))
    return leading // 2
